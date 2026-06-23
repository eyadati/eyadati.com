import subprocess
import sys
import json
import os
import time
import random
import openpyxl
from concurrent.futures import ProcessPoolExecutor, as_completed

CONCURRENCY = 3
OUT_PATH = "/home/oussama/Desktop/eyadati/eyadati/docs/clinics_with_phones.xlsx"
SCRIPT = "/home/oussama/Desktop/eyadati/eyadati/scripts/scrape_one.py"
INPUT_PATH = "/home/oussama/Desktop/eyadati/eyadati/docs/clinics.xlsx"

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_5 like Mac OS X) AppleWebKit/605.1.15 Mobile/15E148",
]

# Non-clinic keywords to skip entirely
SKIP_KEYWORDS = [
    "chu ", "centre hospitalier", "polyclinique", "ecole", "école", "paramedicale",
    "pharmacie", "wilaya", "direction de la santé", "dsp ", "mairie", "hopital",
    "hôpital", "laboratoire", "cabinet d'avocat", "audi", "coiffure", "beauty",
    "aicha la japonaise", "imama", "ferradj", "khedim", "bensalem", "tlemcen",
    "renadial", "ca ", "c a", "audiol", "cabinet dentaire"  # skip generic "cabinet dentaire" entries
]

def should_skip(name):
    name_lower = name.lower()
    for kw in SKIP_KEYWORDS:
        if kw in name_lower:
            return True
    # skip very short / generic names
    if len(name) < 8:
        return True
    return False

def save_result(name, phone, rating, url):
    try:
        wb = openpyxl.load_workbook(OUT_PATH)
        ws = wb.active
        ws.append([name, phone, str(rating) if rating else "", url])
        wb.save(OUT_PATH)
        return True
    except Exception as e:
        with open(OUT_PATH + ".csv", "a") as f:
            f.write(f"{name}\t{phone}\t{rating}\t{url}\n")
        return False

def scrape_one_with_ua(name, url):
    """Run subprocess with a random user agent injected."""
    ua = random.choice(USER_AGENTS)
    try:
        result = subprocess.run(
            [sys.executable, SCRIPT, name, url],
            capture_output=True, text=True, timeout=45,
            env={**os.environ, "PLAYWRIGHT_UA": ua}
        )
        if result.returncode != 0:
            return {"name": name, "phone": "", "error": f"exit {result.returncode}"}
        for line in reversed(result.stdout.strip().split("\n")):
            line = line.strip()
            if line.startswith("{"):
                return json.loads(line)
        return {"name": name, "phone": "", "error": "no JSON"}
    except subprocess.TimeoutExpired:
        return {"name": name, "phone": "", "error": "timeout"}
    except Exception as e:
        return {"name": name, "phone": "", "error": str(e)[:80]}

def main():
    # Read all names from output (those already with phones)
    done_names = set()
    if os.path.exists(OUT_PATH):
        try:
            wb_out = openpyxl.load_workbook(OUT_PATH, data_only=True)
            for row in wb_out.active.iter_rows(values_only=True):
                name = str(row[0]).strip() if row[0] else ""
                if name and name != "Name":
                    done_names.add(name)
        except:
            pass

    # Read input xlsx to get todo list
    wb_in = openpyxl.load_workbook(INPUT_PATH, read_only=True, data_only=True)
    ws_in = wb_in.active
    todo = []
    for row in ws_in.iter_rows(values_only=True):
        if row[0] and row[1]:
            name = str(row[1]).strip()
            url = str(row[0]).strip()
            rating = str(row[2]) if row[2] else ""
            if name and url and name not in done_names:
                todo.append((name, url, rating))

    total = len(todo)
    print(f"Output has {len(done_names)} with phones, {total} remaining to scrape")
    if total == 0:
        print("Nothing to scrape!")
        return

    success = 0
    start = time.time()

    with ProcessPoolExecutor(max_workers=CONCURRENCY) as executor:
        futures = {}
        for idx, (name, url, rating) in enumerate(todo, start=1):
            future = executor.submit(scrape_one_with_ua, name, url)
            futures[future] = (idx, name, rating, url)
            # Random delay between launches to avoid rate limiting
            time.sleep(random.uniform(1.0, 3.0))

        for future in as_completed(futures):
            idx, name, rating, url = futures[future]
            result = future.result()
            phone = result.get("phone", "")
            err = result.get("error", "")

            if phone:
                save_result(name, phone, rating, url)
                success += 1
                print(f"[{idx}/{total}] OK  {name[:45]:45s} → {phone}")
            else:
                if err:
                    print(f"[{idx}/{total}] ERR {name[:40]:40s} → {err}")
                else:
                    print(f"[{idx}/{total}] --  {name[:45]:45s} → no phone")

            # Random delay after each result
            time.sleep(random.uniform(1.0, 2.0))

    elapsed = time.time() - start
    print(f"\n{'='*50}")
    print(f"Done in {elapsed/60:.1f} min")
    print(f"New successes: {success}/{total}")

if __name__ == "__main__":
    main()
