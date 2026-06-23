import subprocess
import sys
import json
import os
import time
import openpyxl
from concurrent.futures import ProcessPoolExecutor, as_completed

CONCURRENCY = 4
OUT_PATH = "/home/oussama/Desktop/eyadati/eyadati/docs/clinics_with_phones.xlsx"
SCRIPT = "/home/oussama/Desktop/eyadati/eyadati/scripts/scrape_one.py"

def save_result(name, phone, rating, url):
    try:
        if not os.path.exists(OUT_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clinics with Phones"
            ws.append(["Name", "Phone", "Rating", "URL"])
        else:
            wb = openpyxl.load_workbook(OUT_PATH)
            ws = wb.active
        ws.append([name, phone, str(rating) if rating else "", url])
        wb.save(OUT_PATH)
        return True
    except Exception as e:
        # fallback CSV
        with open(OUT_PATH + ".csv", "a") as f:
            f.write(f"{name}\t{phone}\t{rating}\t{url}\n")
        return False

def scrape_one(name, url):
    """Run in a subprocess so Playwright Node crashes don't kill us."""
    try:
        result = subprocess.run(
            [sys.executable, SCRIPT, name, url],
            capture_output=True, text=True, timeout=45
        )
        if result.returncode != 0:
            return {"name": name, "phone": "", "error": f"exit {result.returncode}"}
        # last line should be JSON
        for line in reversed(result.stdout.strip().split("\n")):
            line = line.strip()
            if line.startswith("{"):
                return json.loads(line)
        return {"name": name, "phone": "", "error": "no JSON output"}
    except subprocess.TimeoutExpired:
        return {"name": name, "phone": "", "error": "timeout"}
    except Exception as e:
        return {"name": name, "phone": "", "error": str(e)[:80]}

def main():
    # Read input xlsx
    wb_in = openpyxl.load_workbook(
        "/home/oussama/Desktop/eyadati/eyadati/docs/clinics.xlsx",
        read_only=True, data_only=True,
    )
    ws_in = wb_in.active
    all_rows = []
    for i, row in enumerate(ws_in.iter_rows(values_only=True)):
        if i == 0:
            continue
        url, name, rating = row[0], row[1], str(row[2]) if row[2] else ""
        if url and name:
            all_rows.append((str(url).strip(), str(name).strip(), rating))

    # Read already-done names from output file
    done_names = set()
    if os.path.exists(OUT_PATH):
        try:
            wb_out = openpyxl.load_workbook(OUT_PATH, data_only=True)
            ws_out = wb_out.active
            for row in ws_out.iter_rows(values_only=True):
                if row[0] and row[0] != "Name":
                    done_names.add(row[0].strip())
        except:
            pass

    todo = [(u, n, r) for (u, n, r) in all_rows if n not in done_names]
    total = len(todo)
    already = len(all_rows) - total
    print(f"Already saved: {already}, Remaining: {total}")

    if total == 0:
        print("All done!")
        return

    start = time.time()
    success = 0

    with ProcessPoolExecutor(max_workers=CONCURRENCY) as executor:
        futures = {}
        for idx, (url, name, rating) in enumerate(todo, start=1):
            future = executor.submit(scrape_one, name, url)
            futures[future] = (idx, name, rating, url)

        for future in as_completed(futures):
            idx, name, rating, url = futures[future]
            result = future.result()
            phone = result.get("phone", "")
            err = result.get("error", "")

            if phone:
                save_result(name, phone, rating, url)
                success += 1
                print(f"[{idx}/{total}] OK  {name[:45]:45s} → {phone}")
            else:
                if err:
                    print(f"[{idx}/{total}] ERR {name[:40]:40s} → {err}")
                else:
                    print(f"[{idx}/{total}] --  {name[:45]:45s} → no phone")

    elapsed = time.time() - start
    print(f"\n{'='*50}")
    print(f"Done in {elapsed/60:.1f} min")
    print(f"Success: {success}/{total} (total in file: {already + success})")

if __name__ == "__main__":
    main()
