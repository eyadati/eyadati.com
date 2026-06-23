import asyncio
import re
import time
import os
import openpyxl
from playwright.async_api import async_playwright, TimeoutError

CONCURRENCY = 4
PAGE_TIMEOUT = 20000
STAGGER_DELAY = 1.5

DZ_PHONE_STRICT = re.compile(r'(0[567]\d{8}|\+213[567]\d{8}|00213[567]\d{8})')
DZ_PHONE_RE = re.compile(r'(?:(?:\+213|00213)|\b0)(5|6|7)\s*(\d{2})\s*(\d{2})\s*(\d{2})\s*(\d{2})')

SEMAPHORE = asyncio.Semaphore(CONCURRENCY)
OUT_PATH = "/home/oussama/Desktop/eyadati/eyadati/docs/clinics_with_phones.xlsx"
lock = asyncio.Lock()

HEADERS = {"Accept-Language": "fr-FR,fr;q=0.9"}

def normalize(raw: str) -> str:
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 10 and digits.startswith('0'):
        return '+213' + digits[1:]
    if len(digits) == 12 and digits.startswith('213'):
        return '+' + digits
    return raw.strip()

def is_phone(s: str) -> bool:
    return bool(re.match(r'^\+213[567]\d{8}$', s))

def save_result(name, phone, rating, url):
    """Append a row to the xlsx file (thread-safe via lock)."""
    try:
        if not os.path.exists(OUT_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clinics with Phones"
            ws.append(["Name", "Phone", "Rating", "URL"])
        else:
            wb = openpyxl.load_workbook(OUT_PATH)
            ws = wb.active
        ws.append([name, phone, str(rating) if rating else "", url])
        wb.save(OUT_PATH)
    except Exception as e:
        # fallback: append via CSV-like method if xlsx fails
        with open(OUT_PATH + ".csv", "a") as f:
            f.write(f"{name}\t{phone}\t{rating}\t{url}\n")

async def extract_phone(page):
    """Try multiple strategies to extract a phone number."""
    phones = set()

    # Strategy 1: tel: links
    links = await page.query_selector_all('a[href*="tel:"]')
    for a in links:
        href = await a.get_attribute("href")
        if href:
            num = href.replace("tel:", "").split("?")[0].split("&")[0]
            for m in DZ_PHONE_STRICT.finditer(num):
                p = m.group(0)
                if len(p) >= 9:
                    phones.add(p)

    # Strategy 2: wait for Google Maps phone element
    try:
        el = await page.wait_for_selector(
            '[data-item-id="phone"] button, [data-tooltip*="Copier"], button[aria-label*="téléphone"], button[aria-label*="phone"]',
            timeout=8000
        )
        text = await el.inner_text()
        for m in DZ_PHONE_STRICT.finditer(text):
            phones.add(m.group(0))
        for attr in ["data-tooltip", "aria-label"]:
            val = await el.get_attribute(attr)
            if val:
                for m in DZ_PHONE_STRICT.finditer(val):
                    phones.add(m.group(0))
    except:
        pass

    # Strategy 3: scan all button texts
    buttons = await page.query_selector_all("button")
    for btn in buttons:
        text = await btn.inner_text()
        for m in DZ_PHONE_STRICT.finditer(text):
            phones.add(m.group(0))
        aria = await btn.get_attribute("aria-label")
        if aria:
            for m in DZ_PHONE_STRICT.finditer(aria):
                phones.add(m.group(0))

    # Strategy 4: scan full page text
    body = await page.inner_text("body")
    for m in DZ_PHONE_STRICT.finditer(body):
        p = m.group(0)
        if len(p) >= 10 and re.search(r'[567]', p):
            phones.add(p)

    # Strategy 5: look for phone spans (Google Maps specific)
    spans = await page.query_selector_all('[class*="phone"], [class*="Phone"], span[aria-label*="phone"]')
    for s in spans:
        text = await s.inner_text()
        for m in DZ_PHONE_STRICT.finditer(text):
            phones.add(m.group(0))

    return phones

async def process_url(browser, url, name, rating, idx, total):
    async with SEMAPHORE:
        context = None
        try:
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0.0.0 Safari/537.36",
                locale="fr-FR",
                extra_http_headers=HEADERS,
            )
            page = await context.new_page()
            await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)

            # Let page render
            try:
                await page.wait_for_timeout(3000)
            except:
                pass

            phones = await extract_phone(page)

            # Clean up
            await page.close()
            await context.close()
            context = None

            if phones:
                normalized = []
                for p in phones:
                    n = normalize(p)
                    if is_phone(n) and n not in normalized:
                        normalized.append(n)
                if normalized:
                    phone_str = "; ".join(normalized)
                    print(f"[{idx}/{total}] OK  {name[:45]:45s} → {normalized[0]}")
                    async with lock:
                        save_result(name, phone_str, rating, url)
                    return True
                else:
                    print(f"[{idx}/{total}] --  {name[:45]:45s} → invalid phone format")
                    return False
            else:
                print(f"[{idx}/{total}] --  {name[:45]:45s} → no phone")
                return False

        except TimeoutError:
            print(f"[{idx}/{total}] TIMEOUT {name[:40]:40s}")
            return False
        except Exception as e:
            err = str(e)[:60]
            print(f"[{idx}/{total}] ERR {name[:40]:40s} → {err}")
            return False
        finally:
            if context:
                try:
                    await context.close()
                except:
                    pass

async def main():
    wb = openpyxl.load_workbook(
        "/home/oussama/Desktop/eyadati/eyadati/docs/clinics.xlsx",
        read_only=True, data_only=True,
    )
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        url, name, rating = row[0], row[1], str(row[2]) if row[2] else ""
        if url and name:
            rows.append((str(url).strip(), str(name).strip(), rating))

    total = len(rows)
    print(f"Total: {total} URLs (concurrency={CONCURRENCY})\n")

    start = time.time()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
        )

        sem = asyncio.Semaphore(CONCURRENCY)

        async def worker(url, name, rating, idx):
            async with sem:
                return await process_url(browser, url, name, rating, idx, total)

        tasks = []
        for idx, (url, name, rating) in enumerate(rows, start=1):
            tasks.append(asyncio.create_task(worker(url, name, rating, idx)))
            await asyncio.sleep(STAGGER_DELAY / CONCURRENCY)

        await asyncio.gather(*tasks, return_exceptions=True)
        await browser.close()

    elapsed = time.time() - start
    print(f"\nDone in {elapsed/60:.1f} min")

if __name__ == "__main__":
    asyncio.run(main())
