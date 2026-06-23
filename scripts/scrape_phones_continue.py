import asyncio
import re
import os
import time
import openpyxl
from playwright.async_api import async_playwright, TimeoutError

CONCURRENCY = 3
PAGE_TIMEOUT = 20000
STAGGER_DELAY = 2.0

DZ_PHONE_STRICT = re.compile(r'(0[567]\d{8}|\+213[567]\d{8}|00213[567]\d{8})')
OUT_PATH = "/home/oussama/Desktop/eyadati/eyadati/docs/clinics_with_phones.xlsx"

def normalize(raw: str) -> str:
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 10 and digits.startswith('0'):
        return '+213' + digits[1:]
    if len(digits) == 12 and digits.startswith('213'):
        return '+' + digits
    return raw.strip()

def is_phone(s: str) -> bool:
    return bool(re.match(r'^\+213[567]\d{8}$', s))

def save_result(name, phone, rating, url):
    try:
        if not os.path.exists(OUT_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clinics with Phones"
            ws.append(["Name", "Phone", "Rating", "URL"])
        else:
            wb = openpyxl.load_workbook(OUT_PATH)
            ws = wb.active
        ws.append([name, phone, str(rating) if rating else "", url])
        wb.save(OUT_PATH)
    except:
        with open(OUT_PATH + ".csv", "a") as f:
            f.write(f"{name}\t{phone}\t{rating}\t{url}\n")

async def extract_phone(page):
    phones = set()
    try:
        links = await page.query_selector_all('a[href*="tel:"]')
        for a in links:
            href = await a.get_attribute("href")
            if href:
                num = href.replace("tel:", "").split("?")[0].split("&")[0]
                for m in DZ_PHONE_STRICT.finditer(num):
                    phones.add(m.group(0))
    except:
        pass
    try:
        el = await page.wait_for_selector(
            '[data-item-id="phone"] button, [data-tooltip*="Copier"], button[aria-label*="téléphone"]',
            timeout=6000
        )
        text = await el.inner_text()
        for m in DZ_PHONE_STRICT.finditer(text):
            phones.add(m.group(0))
        for attr in ["data-tooltip", "aria-label"]:
            val = await el.get_attribute(attr)
            if val:
                for m in DZ_PHONE_STRICT.finditer(val):
                    phones.add(m.group(0))
    except:
        pass
    try:
        for btn in await page.query_selector_all("button"):
            text = await btn.inner_text()
            for m in DZ_PHONE_STRICT.finditer(text):
                phones.add(m.group(0))
            aria = await btn.get_attribute("aria-label")
            if aria:
                for m in DZ_PHONE_STRICT.finditer(aria):
                    phones.add(m.group(0))
    except:
        pass
    try:
        body = await page.inner_text("body")
        for m in DZ_PHONE_STRICT.finditer(body):
            p = m.group(0)
            if len(p) >= 10:
                phones.add(p)
    except:
        pass
    return phones

async def process_one(url, name, rating, idx, total):
    """Process a single URL with its own browser instance (prevents EPIPE crashes)."""
    async with async_playwright() as p:
        browser = None
        try:
            browser = await p.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
            )
            ctx = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0.0.0 Safari/537.36",
                locale="fr-FR",
            )
            page = await ctx.new_page()
            await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
            await page.wait_for_timeout(3000)
            phones = await extract_phone(page)
            await page.close()
            await ctx.close()

            if phones:
                normalized = []
                for pn in phones:
                    n = normalize(pn)
                    if is_phone(n) and n not in normalized:
                        normalized.append(n)
                if normalized:
                    phone_str = "; ".join(normalized)
                    print(f"[{idx}/{total}] OK  {name[:45]:45s} → {normalized[0]}")
                    save_result(name, phone_str, rating, url)
                    return True
            print(f"[{idx}/{total}] --  {name[:45]:45s} → no phone")
            return False

        except Exception as e:
            err = str(e)[:60]
            print(f"[{idx}/{total}] ERR {name[:40]:40s} → {err}")
            return False
        finally:
            if browser:
                try:
                    await browser.close()
                except:
                    pass

async def main():
    # Read input
    wb_in = openpyxl.load_workbook(
        "/home/oussama/Desktop/eyadati/eyadati/docs/clinics.xlsx",
        read_only=True, data_only=True,
    )
    ws_in = wb_in.active
    all_rows = []
    for i, row in enumerate(ws_in.iter_rows(values_only=True)):
        if i == 0:
            continue
        url, name, rating = row[0], row[1], str(row[2]) if row[2] else ""
        if url and name:
            all_rows.append((str(url).strip(), str(name).strip(), rating))

    # Read already-done
    done_names = set()
    if os.path.exists(OUT_PATH):
        wb_out = openpyxl.load_workbook(OUT_PATH, data_only=True)
        ws_out = wb_out.active
        for row in ws_out.iter_rows(values_only=True):
            if row[0] and row[0] != "Name":
                done_names.add(row[0].strip())

    todo = [(u, n, r) for (u, n, r) in all_rows if n not in done_names]
    total = len(todo)
    print(f"Already saved: {len(all_rows) - total}, Remaining: {total}")

    if total == 0:
        print("All done!")
        return

    start = time.time()
    sem = asyncio.Semaphore(CONCURRENCY)

    async def worker(url, name, rating, idx):
        async with sem:
            return await process_one(url, name, rating, idx, total)

    tasks = []
    for idx, (url, name, rating) in enumerate(todo, start=1):
        tasks.append(asyncio.create_task(worker(url, name, rating, idx)))
        await asyncio.sleep(STAGGER_DELAY / CONCURRENCY)

    results = await asyncio.gather(*tasks, return_exceptions=True)
    success = sum(1 for r in results if r is True)
    elapsed = time.time() - start

    print(f"\n{'='*50}")
    print(f"Done in {elapsed/60:.1f} min")
    print(f"Success: {success}/{total} (total in file: {len(all_rows) - total + success})")

if __name__ == "__main__":
    asyncio.run(main())
