import openpyxl
import re
import time
import json
import sys
from urllib.parse import urlparse, parse_qs
import requests
from bs4 import BeautifulSoup

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept-Language": "fr-FR,fr;q=0.9",
}

ALGERIAN_PHONE_RE = re.compile(
    r"(?:\+213|00213|0)(5|6|7)\s*\d{2}\s*\d{2}\s*\d{2}\s*\d{2}"
)
GENERAL_PHONE_RE = re.compile(r"\+?[\d\s\-\(\)]{7,20}")

def extract_place_id(url):
    """Extract Google Place ID from a Maps URL."""
    m = re.search(r"1s([a-zA-Z0-9:_\-]+)!", url)
    if m:
        return m.group(1)
    m = re.search(r"place/[^/]+/[^/]+ChIJ([^/?&]+)", url)
    if m:
        return "ChIJ" + m.group(1)
    m = re.search(r"ChIJ[^/?&]+", url)
    if m:
        return m.group(0)
    return None

def fetch_page(url):
    """Fetch a Google Maps page and return HTML text, or None."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code == 200:
            return resp.text
        return None
    except Exception as e:
        return None

def find_phones_in_html(html, url):
    """Find phone numbers in HTML using multiple strategies."""
    phones = set()
    soup = BeautifulSoup(html, "lxml")

    # Strategy 1: JSON-LD structured data
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string)
            if isinstance(data, dict):
                for key in ["telephone", "phone", "contactPoint"]:
                    val = data.get(key)
                    if isinstance(val, str) and re.search(r"\d", val):
                        phones.add(val)
                    elif isinstance(val, dict) and "telephone" in val:
                        phones.add(val["telephone"])
            elif isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        for key in ["telephone", "phone"]:
                            val = item.get(key)
                            if isinstance(val, str) and re.search(r"\d", val):
                                phones.add(val)
        except:
            pass

    # Strategy 2: Meta tags
    for meta in soup.find_all("meta"):
        content = meta.get("content", "")
        if re.search(r"\+?0(5|6|7)\d", content):
            phones.add(content.strip())

    # Strategy 3: Search for phone patterns in text
    for tag in soup.find_all(["span", "div", "a", "button"]):
        text = tag.get_text(strip=True)
        for match in ALGERIAN_PHONE_RE.findall(text):
            full = ALGERIAN_PHONE_RE.search(text)
            if full:
                phones.add(full.group(0).strip())
        # Also find any text node containing "phone" or "tel"
        for attr in ["href", "aria-label", "title"]:
            val = tag.get(attr, "")
            if "tel:" in val:
                num = val.replace("tel:", "").split("?")[0]
                phones.add(num)

    return phones

def normalize_phone(phone_str):
    """Normalize to E.164 if possible."""
    digits = re.sub(r"\D", "", phone_str)
    if len(digits) >= 9:
        if len(digits) == 9:
            digits = "213" + digits
        if len(digits) == 10 and digits.startswith("0"):
            digits = "213" + digits[1:]
        if len(digits) == 12 and digits.startswith("213"):
            return "+" + digits
        if len(digits) == 13 and digits.startswith("0213"):
            return "+" + digits[1:]
    return phone_str.strip()

def process_row(url, name, rating, idx):
    """Process a single row, return result dict or None."""
    place_id = extract_place_id(url)
    html = fetch_page(url)
    if not html:
        return None

    raw_phones = find_phones_in_html(html, url)
    if not raw_phones:
        # try without trailing params
        clean_url = url.split("?")[0]
        if clean_url != url:
            html2 = fetch_page(clean_url)
            if html2:
                raw_phones = find_phones_in_html(html2, clean_url)

    if raw_phones:
        normalized = [normalize_phone(p) for p in raw_phones]
        unique = list(dict.fromkeys(normalized))
        # filter to keep only plausible phone strings
        valid = [p for p in unique if re.search(r"\d{6,}", p)]
        if valid:
            return {
                "name": name,
                "rating": rating,
                "phone": "; ".join(valid),
                "place_id": place_id or "",
                "url": url,
            }
    return None

def main():
    wb = openpyxl.load_workbook(
        "/home/oussama/Desktop/eyadati/eyadati/docs/clinics.xlsx",
        read_only=True,
        data_only=True,
    )
    ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        url, name, rating = row[0], row[1], row[2]
        if not url or not name:
            continue
        rows.append((str(url), str(name), rating))

    print(f"Total rows to process: {len(rows)}", flush=True)

    results = []
    success_count = 0

    for idx, (url, name, rating) in enumerate(rows):
        print(f"[{idx+1}/{len(rows)}] {name[:40]}...", end=" ", flush=True)
        result = process_row(url, name, rating, idx)
        if result:
            success_count += 1
            results.append(result)
            print(f"PHONE: {result['phone']}", flush=True)
        else:
            print("no phone", flush=True)
        time.sleep(1.5)  # be gentle with Google

    print(f"\nDone. {success_count}/{len(rows)} successes.")

    # Write results to new xlsx
    out_path = "/home/oussama/Desktop/eyadati/eyadati/docs/clinics_with_phones.xlsx"
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Clinics with Phones"
    out_ws.append(["Name", "Phone", "Rating", "Place ID", "URL"])
    for r in results:
        out_ws.append([r["name"], r["phone"], r["rating"], r["place_id"], r["url"]])

    out_wb.save(out_path)
    print(f"Saved results to {out_path}")

if __name__ == "__main__":
    main()
